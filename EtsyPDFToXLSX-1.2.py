import subprocess
import importlib
import sys

def install_and_import(package, import_name=None):
    if import_name is None:
        import_name = package
    try:
        importlib.import_module(import_name)
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", package])
    finally:
        globals()[import_name] = importlib.import_module(import_name)

# Correcting the package list with the proper pip-installable names
packages = [('PyMuPDF', 'fitz'), 'openpyxl', 'tkinter']
for package in packages:
    if isinstance(package, tuple):
        install_and_import(*package)
    else:
        install_and_import(package)
# Import other necessary modules after package checks
import tkinter as tk
from tkinter import filedialog, messagebox
import fitz  # PyMuPDF
import re
from openpyxl import Workbook

class PDFConverterApp:
    def __init__(self, master):
        self.master = master
        master.title("PDF to Excel Converter")

        self.select_pdf_button = tk.Button(master, text="Select PDF File", command=self.select_pdf)
        self.select_pdf_button.pack()

        self.select_save_location_button = tk.Button(master, text="Select Save Location", command=self.select_save_location)
        self.select_save_location_button.pack()

        self.convert_button = tk.Button(master, text="Start Conversion", command=self.start_conversion)
        self.convert_button.pack()

        self.pdf_path = ""
        self.save_path = ""

    def select_pdf(self):
        self.pdf_path = filedialog.askopenfilename(filetypes=[("PDF files", "*.pdf")])

    def select_save_location(self):
        self.save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])

    def start_conversion(self):
        if not self.pdf_path or not self.save_path:
            messagebox.showerror("Error", "Please select both a PDF file and a save location.")
            return

        try:
            doc = fitz.open(self.pdf_path)
            wb = Workbook()
            
            # Create separate sheets
            ws_skins = wb.create_sheet("Credit Card Skins")
            ws_skins.append(["Order Number", "SKU", "Chip Size", "Finish", "Quantity"])
            
            ws_stickers = wb.create_sheet("Stickers")
            ws_stickers.append(["Order Number", "SKU", "Size", "Quantity"])
            
            # Remove the default sheet
            wb.remove(wb['Sheet'])

            order_number_pattern = re.compile(r"Order #(\d+)")
            credit_card_item_pattern = re.compile(r"(SKU: CRD-SKN-\d+)\nChip Size: (\w+ \w+)\nFinish: (\w+)\n(\d+) x")
            sticker_item_pattern = re.compile(r"(SKU: STK-\d+)\nSize: (\w+ \w+)\n(\d+) x")

            for page in doc:
                text = page.get_text("text")
                if order_matches := order_number_pattern.search(text):
                    current_order = order_matches.group(1)
                    
                credit_card_items = credit_card_item_pattern.findall(text)
                sticker_items = sticker_item_pattern.findall(text)

                for item in credit_card_items:
                    ws_skins.append([current_order] + list(item))
                for item in sticker_items:
                    ws_stickers.append([current_order] + list(item))

            wb.save(self.save_path)
            messagebox.showinfo("Success", "Conversion successful! File saved.")
            
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")

root = tk.Tk()
app = PDFConverterApp(root)
root.mainloop()


